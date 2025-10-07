#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表生成模块
提供清单生成、Excel导出、打印等功能
"""

import os
import json
import io
import base64
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

# Excel相关库
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from database import Database

class ReportGenerator:
    """报表生成器"""
    
    def __init__(self, db_manager: Database):
        """初始化报表生成器"""
        self.db = db_manager
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
        
        # 报表模板配置
        self.report_config = {
            'company_name': '哈迪斯',
            'report_title': '包裹打托清单',
            'font_size': 12,
            'header_font_size': 16,
            'margin': 20
        }
    
    def _get_package_type_display(self, package: Dict) -> str:
        """获取包裹类型的显示文本"""
        if package['package_type'] == 'existing':
            return "板件"
        elif package['package_type'] == 'custom':
            return package.get('custom_type_name') or "自定义"
        else:
            return "未知"
    
    def generate_print_report(self, tray_code: str) -> str:
        """生成打印报告"""
        try:
            # 获取托盘信息
            tray_info = self.db.get_tray_info(tray_code)
            if not tray_info:
                raise ValueError(f"托盘 {tray_code} 不存在")
            
            # 获取包裹列表
            packages = self.db.get_tray_packages(tray_code)
            
            # 生成HTML格式的打印报告
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"print_report_{tray_code}_{timestamp}.html"
            file_path = self.reports_dir / filename
            
            html_content = self._generate_print_html(tray_info, packages)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return str(file_path)
            
        except Exception as e:
            print(f"生成打印报告错误: {e}")
            return None
    
    def _generate_print_html(self, tray_info: Dict, packages: List[Dict]) -> str:
        """生成打印用的HTML内容"""
        html = f'''
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>托盘清单 - {tray_info['tray_code']}</title>
    <style>
        @media print {{
            body {{ margin: 0; }}
            .no-print {{ display: none; }}
        }}
        body {{
            font-family: Arial, sans-serif;
            margin: 20px;
            font-size: 12px;
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
            border-bottom: 2px solid #333;
            padding-bottom: 10px;
        }}
        .company-name {{
            font-size: 18px;
            font-weight: bold;
            margin-bottom: 5px;
        }}
        .report-title {{
            font-size: 16px;
            margin-bottom: 10px;
        }}
        .tray-info {{
            margin-bottom: 20px;
            background-color: #f5f5f5;
            padding: 10px;
            border-radius: 5px;
        }}
        .info-row {{
            margin-bottom: 5px;
        }}
        .label {{
            font-weight: bold;
            display: inline-block;
            width: 100px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }}
        th, td {{
            border: 1px solid #333;
            padding: 8px;
            text-align: left;
            vertical-align: middle;
        }}
        th {{
            background-color: #f0f0f0;
            font-weight: bold;
        }}
        .footer {{
            margin-top: 30px;
            text-align: center;
            font-size: 10px;
            color: #666;
        }}
    </style>
</head>
<body>
    <div class="header">
        <div class="company-name">{self.report_config['company_name']}</div>
        <div class="report-title">{self.report_config['report_title']}</div>
    </div>
    
    <div class="tray-info">
        <div class="info-row">
            <span class="label">托盘编号:</span>
            <span>{tray_info['tray_code']}</span>
        </div>
        <div class="info-row">
            <span class="label">托盘类型:</span>
            <span>{'实体托盘' if tray_info['tray_type'] == 'physical' else '虚拟托盘'}</span>
        </div>
        <div class="info-row">
            <span class="label">容量限制:</span>
            <span>{tray_info['max_capacity']} 件</span>
        </div>
        <div class="info-row">
            <span class="label">当前数量:</span>
            <span>{tray_info['current_count']} 件</span>
        </div>
        <div class="info-row">
            <span class="label">创建时间:</span>
            <span>{tray_info['created_at']}</span>
        </div>
        <div class="info-row">
            <span class="label">打印时间:</span>
            <span>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</span>
        </div>
    </div>
    
    <table>
        <thead>
            <tr>
                <th>序号</th>
                <th>包裹号</th>
                <th>包裹类型</th>
                <th>状态</th>
                <th>创建时间</th>
            </tr>
        </thead>
        <tbody>
'''
        
        for i, package in enumerate(packages, 1):
            html += f'''
            <tr>
                <td>{i}</td>
                <td>{package['package_code']}</td>
                <td>{self._get_package_type_display(package)}</td>
                <td>{package['status']}</td>
                <td>{package['created_at']}</td>
            </tr>
'''
        
        html += f'''
        </tbody>
    </table>
    
    <div class="footer">
        <p>共 {len(packages)} 个包裹 | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
</body>
</html>
'''
        return html
    
    def export_to_excel(self, tray_code: str = None, packages: List[Dict] = None) -> str:
        """导出到Excel"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl库未安装，无法导出Excel")
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if tray_code:
                ws.title = f"托盘_{tray_code}"
                # 获取托盘信息
                tray_info = self.db.get_tray_info(tray_code)
                packages = self.db.get_tray_packages(tray_code)
                
                # 设置标题
                ws['A1'] = f"{self.report_config['company_name']} - {self.report_config['report_title']}"
                ws.merge_cells('A1:E1')
                ws['A1'].font = Font(size=16, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center')
                
                # 托盘信息
                ws['A3'] = f"托盘编号: {tray_info['tray_code']}"
                ws['A4'] = f"托盘类型: {'实体托盘' if tray_info['tray_type'] == 'physical' else '虚拟托盘'}"
                ws['A5'] = f"容量: {tray_info['current_count']}/{tray_info['max_capacity']}"
                
                # 表头
                headers = ['序号', '包裹号', '包裹类型', '状态', '创建时间']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=7, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # 数据行
                for row, package in enumerate(packages, 8):
                    ws.cell(row=row, column=1, value=row-7)
                    ws.cell(row=row, column=2, value=package['package_code'])
                    ws.cell(row=row, column=3, value=self._get_package_type_display(package))
                    ws.cell(row=row, column=4, value=package['status'])
                    ws.cell(row=row, column=5, value=package['created_at'])
                
                filename = f"tray_report_{tray_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            else:
                ws.title = "包裹清单"
                # 设置标题
                ws['A1'] = f"{self.report_config['company_name']} - 包裹清单"
                ws.merge_cells('A1:E1')
                ws['A1'].font = Font(size=16, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center')
                
                # 表头
                headers = ['序号', '包裹号', '包裹类型', '托盘', '状态', '创建时间']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # 数据行
                for row, package in enumerate(packages, 4):
                    ws.cell(row=row, column=1, value=row-3)
                    ws.cell(row=row, column=2, value=package['package_code'])
                    ws.cell(row=row, column=3, value=self._get_package_type_display(package))
                    ws.cell(row=row, column=4, value=package.get('tray_id', ''))
                    ws.cell(row=row, column=5, value=package['status'])
                    ws.cell(row=row, column=6, value=package['created_at'])
                
                filename = f"package_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            # 调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            file_path = self.reports_dir / filename
            wb.save(file_path)
            
            return str(file_path)
            
        except Exception as e:
            print(f"导出Excel错误: {e}")
            return None
    
    def generate_summary_report(self) -> Dict[str, Any]:
        """生成汇总报告"""
        try:
            # 获取所有托盘
            trays = self.db.get_all_trays()
            
            # 获取所有包裹
            packages = self.db.search_packages("")
            
            # 统计信息
            total_trays = len(trays)
            total_packages = len(packages)
            
            # 按状态统计托盘
            tray_status_stats = {}
            for tray in trays:
                status = tray['status']
                tray_status_stats[status] = tray_status_stats.get(status, 0) + 1
            
            # 按类型统计托盘
            tray_type_stats = {}
            for tray in trays:
                tray_type = tray['tray_type']
                tray_type_stats[tray_type] = tray_type_stats.get(tray_type, 0) + 1
            
            # 按状态统计包裹
            package_status_stats = {}
            for package in packages:
                status = package['status']
                package_status_stats[status] = package_status_stats.get(status, 0) + 1
            
            # 按类型统计包裹
            package_type_stats = {}
            for package in packages:
                pkg_type = self._get_package_type_display(package)
                package_type_stats[pkg_type] = package_type_stats.get(pkg_type, 0) + 1
            
            return {
                'total_trays': total_trays,
                'total_packages': total_packages,
                'tray_status_stats': tray_status_stats,
                'tray_type_stats': tray_type_stats,
                'package_status_stats': package_status_stats,
                'package_type_stats': package_type_stats,
                'generated_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            print(f"生成汇总报告错误: {e}")
            return {}
    
    def save_summary_report(self) -> str:
        """保存汇总报告到文件"""
        try:
            summary = self.generate_summary_report()
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"summary_report_{timestamp}.json"
            file_path = self.reports_dir / filename
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
            
            return str(file_path)
            
        except Exception as e:
            print(f"保存汇总报告错误: {e}")
            return None