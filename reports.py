import sys
import os
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
                             QPushButton, QTableWidget, QTableWidgetItem, QLabel,
                             QLineEdit, QTextEdit, QComboBox, QMessageBox,
                             QDialog, QDialogButtonBox, QGroupBox, QCheckBox,
                             QDateEdit, QTabWidget, QHeaderView, QFileDialog,
                             QProgressDialog, QApplication, QCompleter)
from PyQt5.QtCore import Qt, QDate, QThread, pyqtSignal, QStringListModel
from PyQt5.QtGui import QFont, QPainter, QColor
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from qr_handler import QRCodeHandler
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from database import db
from order_management import OrderSelectionDialog

class ExportWorker(QThread):
    """导出工作线程"""
    
    progress_updated = pyqtSignal(int)
    status_updated = pyqtSignal(str)
    export_finished = pyqtSignal(bool, str)
    
    def __init__(self, export_type, file_path, filters=None):
        super().__init__()
        self.export_type = export_type
        self.file_path = file_path
        self.filters = filters or {}
    
    def run(self):
        try:
            if self.export_type == "packages":
                self.export_packages()
            elif self.export_type == "pallets":
                self.export_pallets()
            elif self.export_type == "package_components":
                self.export_package_components()
            elif self.export_type == "pallet_package_components":
                self.export_pallet_package_components()
            elif self.export_type == "comprehensive":
                self.export_comprehensive()
            
            self.export_finished.emit(True, "导出成功")
            
        except Exception as e:
            self.export_finished.emit(False, f"导出失败：{str(e)}")
    
    def export_packages(self):
        """导出包裹数据"""
        self.status_updated.emit("正在查询包裹数据...")
        
        conn = db.get_connection()
        
        # 查询包裹基本信息
        query = '''
            SELECT 
                p.package_number,
                p.order_id,
                o.order_number,
                o.customer_name,
                o.customer_address,
                COUNT(c.id) as component_count,
                p.status,
                p.created_at,
                p.updated_at,
                pal.pallet_number
            FROM packages p
            LEFT JOIN orders o ON p.order_id = o.id
            LEFT JOIN components c ON c.package_id = p.id
            LEFT JOIN pallet_packages pp ON pp.package_id = p.id
            LEFT JOIN pallets pal ON pp.pallet_id = pal.id
            WHERE 1=1
        '''
        
        params = []
        
        # 添加过滤条件
        if self.filters.get('start_date'):
            query += " AND DATE(p.created_at) >= ?"
            params.append(self.filters['start_date'])
        
        if self.filters.get('end_date'):
            query += " AND DATE(p.created_at) <= ?"
            params.append(self.filters['end_date'])
        
        if self.filters.get('order_number'):
            query += " AND o.order_number LIKE ?"
            params.append(f"%{self.filters['order_number']}%")
        
        query += " GROUP BY p.id ORDER BY p.created_at DESC"
        
        df_packages = pd.read_sql_query(query, conn, params=params)
        
        self.progress_updated.emit(30)
        
        # 查询包裹内的板件详情
        self.status_updated.emit("正在查询板件详情...")
        
        component_query = '''
            SELECT 
                p.package_number,
                c.component_name,
                c.material,
                c.finished_size,
                c.component_code,
                c.room_number,
                c.cabinet_number,
                c.scanned_at
            FROM packages p
            JOIN components c ON c.package_id = p.id
            WHERE c.package_id IS NOT NULL
            ORDER BY p.package_number, c.scanned_at
        '''
        
        df_components = pd.read_sql_query(component_query, conn)
        # 结束数据库连接
        conn.close()
        
        self.progress_updated.emit(60)
        
        # 创建Excel文件
        self.status_updated.emit("正在生成Excel文件...")
        
        wb = Workbook()
        
        # 包裹汇总表
        ws_summary = wb.active
        ws_summary.title = "包裹汇总"
        
        # 设置表头
        headers = ['包装号', '订单号', '客户名称', '客户地址', '板件数量', '状态', 
                  '创建时间', '更新时间', '所在托盘']
        
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 填充数据
        for row_idx, row in df_packages.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws_summary.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # 调整列宽
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        self.progress_updated.emit(80)
        
        # 板件详情表
        if not df_components.empty:
            ws_details = wb.create_sheet("板件详情")
            
            detail_headers = ['包装号', '板件名称', '材质', '尺寸', '板件编码', 
                            '房间号', '柜号', '扫描时间']
            
            for col, header in enumerate(detail_headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_idx, row in df_components.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_details.cell(row=row_idx + 2, column=col_idx, value=value)
            
            # 调整列宽
            for column in ws_details.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width
        
        self.progress_updated.emit(90)
        
        # 保存文件
        wb.save(self.file_path)
        self.progress_updated.emit(100)
    
    def export_pallets(self):
        """导出托盘数据（参考用户提供的格式）"""
        self.status_updated.emit("正在查询托盘数据...")
        
        conn = db.get_connection()
        
        # 查询托盘基本信息（统一以 packages.pallet_id 为关联来源）
        query = '''
            SELECT 
                pal.pallet_number,
                pal.pallet_type,
                pal.status,
                pal.created_at,
                pal.sealed_at,
                COUNT(DISTINCT p.id) as package_count,
                COUNT(DISTINCT c.id) as component_count,
                GROUP_CONCAT(DISTINCT o.customer_name) as customers
            FROM pallets pal
            LEFT JOIN packages p ON p.pallet_id = pal.id
            LEFT JOIN components c ON c.package_id = p.id
            LEFT JOIN orders o ON p.order_id = o.id
            WHERE 1=1
        '''
        
        params = []
        
        # 添加过滤条件
        if self.filters.get('start_date'):
            query += " AND DATE(pal.created_at) >= ?"
            params.append(self.filters['start_date'])
        
        if self.filters.get('end_date'):
            query += " AND DATE(pal.created_at) <= ?"
            params.append(self.filters['end_date'])
        
        if self.filters.get('pallet_type'):
            # 将中文类型映射为数据库值
            type_map = {
                '实体托盘': 'physical',
                '虚拟托盘': 'virtual'
            }
            ftype = self.filters['pallet_type']
            db_type = type_map.get(ftype, ftype)
            query += " AND pal.pallet_type = ?"
            params.append(db_type)
        
        query += " GROUP BY pal.id ORDER BY pal.created_at DESC"
        
        df_pallets = pd.read_sql_query(query, conn, params=params)
        
        self.progress_updated.emit(30)
        
        # 查询托盘详细内容
        self.status_updated.emit("正在查询托盘详细内容...")
        
        detail_query = '''
            SELECT 
                pal.pallet_number,
                p.package_number,
                o.order_number,
                o.customer_name,
                COUNT(c.id) as component_count,
                GROUP_CONCAT(c.component_name, '; ') as component_list,
                p.created_at as package_created_at
            FROM pallets pal
            JOIN packages p ON p.pallet_id = pal.id
            LEFT JOIN orders o ON p.order_id = o.id
            LEFT JOIN components c ON c.package_id = p.id
            GROUP BY pal.id, p.id
            ORDER BY pal.pallet_number, p.package_number
        '''
        
        df_details = pd.read_sql_query(detail_query, conn)
        
        # 查询虚拟物品
        virtual_query = '''
            SELECT 
                pal.pallet_number,
                vi.item_name,
                vi.quantity,
                vi.unit,
                vi.remarks,
                vi.created_at
            FROM pallets pal
            JOIN virtual_items vi ON vi.pallet_id = pal.id
            ORDER BY pal.pallet_number, vi.created_at
        '''
        
        df_virtual = pd.read_sql_query(virtual_query, conn)
        conn.close()
        
        self.progress_updated.emit(60)
        
        # 创建Excel文件
        self.status_updated.emit("正在生成Excel文件...")
        
        wb = Workbook()
        
        # 托盘汇总表
        ws_summary = wb.active
        ws_summary.title = "托盘汇总"
        
        # 设置表头（参考用户提供的格式）
        headers = ['托盘号', '托盘类型', '状态', '创建时间', '封托时间', 
                  '包装数量', '板件数量', '客户']
        
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 填充数据
        for row_idx, row in df_pallets.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws_summary.cell(row=row_idx + 2, column=col_idx, value=value)
        
        self.progress_updated.emit(75)
        
        # 托盘详情表
        if not df_details.empty:
            ws_details = wb.create_sheet("托盘详情")
            
            detail_headers = ['托盘号', '包装号', '订单号', '客户名称', '板件数量', 
                            '板件清单', '包装创建时间']
            
            for col, header in enumerate(detail_headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_idx, row in df_details.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_details.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # 虚拟物品表
        if not df_virtual.empty:
            ws_virtual = wb.create_sheet("虚拟物品")
            
            virtual_headers = ['托盘号', '物品名称', '数量', '单位', '备注', '添加时间']
            
            for col, header in enumerate(virtual_headers, 1):
                cell = ws_virtual.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_idx, row in df_virtual.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_virtual.cell(row=row_idx + 2, column=col_idx, value=value)
        
        # 调整所有工作表的列宽
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        self.progress_updated.emit(90)
        
        # 保存文件
        wb.save(self.file_path)
        self.progress_updated.emit(100)
    
    def export_package_components(self):
        """导出包裹-板件明细（满足用户指定字段）"""
        self.status_updated.emit("正在查询包裹-板件明细...")
        conn = db.get_connection()

        query = '''
            SELECT 
                p.package_number AS 包裹号,
                o.order_number AS 订单号,
                c.component_code AS 板件编码,
                c.component_name AS 板件名称,
                c.material AS 材质,
                c.finished_size AS 成品尺寸,
                c.room_number AS 房间,
                c.cabinet_number AS 柜号
            FROM packages p
            JOIN components c ON c.package_id = p.id
            LEFT JOIN orders o ON p.order_id = o.id
            WHERE 1=1
        '''
        params = []

        # 过滤：日期（按包裹创建时间）
        if self.filters.get('start_date'):
            query += " AND DATE(p.created_at) >= ?"
            params.append(self.filters['start_date'])
        if self.filters.get('end_date'):
            query += " AND DATE(p.created_at) <= ?"
            params.append(self.filters['end_date'])

        # 过滤：订单号（模糊）
        if self.filters.get('order_number'):
            query += " AND p.order_id IN (SELECT id FROM orders WHERE order_number LIKE ?)"
            params.append(f"%{self.filters['order_number']}%")

        query += " ORDER BY p.package_number, c.component_code"

        df = pd.read_sql_query(query, conn, params=params)
        # 计算包裹序号（同一订单内按包裹号排序）
        try:
            df.sort_values(by=["订单号","包裹号","板件编码"], inplace=True)
            df["包裹序号"] = (
                df.groupby("订单号")["包裹号"].rank(method="dense").astype(int)
            )
        except Exception:
            pass
        conn.close()

        self.progress_updated.emit(60)

        # 生成Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "包裹-板件明细"

        include_pkg_qr = bool(self.filters.get('include_package_qr'))
        # 表头（使用中文列名）
        base_headers = list(df.columns)
        headers = base_headers + (["包裹二维码"] if include_pkg_qr else [])
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        qr_handler = QRCodeHandler()
        # 数据
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
            if include_pkg_qr:
                try:
                    pkg_num = str(row['包裹号'])
                    img = qr_handler.create_qr_code(pkg_num)
                    bio = BytesIO()
                    img.save(bio, format='PNG')
                    bio.seek(0)
                    xlimg = XLImage(bio)
                    xlimg.width = 80
                    xlimg.height = 80
                    target_col = len(base_headers) + 1
                    ws.add_image(xlimg, ws.cell(row=row_idx + 2, column=target_col).coordinate)
                    ws.row_dimensions[row_idx + 2].height = 60
                    ws.column_dimensions[ws.cell(row=1, column=target_col).column_letter].width = 16
                except Exception:
                    pass

        # 列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        self.progress_updated.emit(90)
        wb.save(self.file_path)
        self.progress_updated.emit(100)

    def export_pallet_package_components(self):
        """导出托盘-包裹-板件明细（可选导出托盘/包裹二维码）"""
        self.status_updated.emit("正在查询托盘-包裹-板件明细...")
        conn = db.get_connection()

        query = '''
            SELECT 
                pal.pallet_number AS 托盘号,
                p.package_number AS 包裹号,
                p.package_index AS 包裹序号,
                o.order_number AS 订单号,
                c.component_code AS 板件编码,
                c.component_name AS 板件名称,
                c.material AS 材质,
                c.finished_size AS 成品尺寸,
                c.room_number AS 房间,
                c.cabinet_number AS 柜号
            FROM pallets pal
            JOIN packages p ON p.pallet_id = pal.id
            JOIN components c ON c.package_id = p.id
            LEFT JOIN orders o ON p.order_id = o.id
            WHERE 1=1
        '''
        params = []
        # 日期过滤（按包裹创建时间）
        if self.filters.get('start_date'):
            query += " AND DATE(p.created_at) >= ?"
            params.append(self.filters['start_date'])
        if self.filters.get('end_date'):
            query += " AND DATE(p.created_at) <= ?"
            params.append(self.filters['end_date'])
        # 订单过滤
        if self.filters.get('order_number'):
            query += " AND o.order_number LIKE ?"
            params.append(f"%{self.filters['order_number']}%")
        # 托盘类型过滤
        if self.filters.get('pallet_type'):
            type_map = {'实体托盘': 'physical', '虚拟托盘': 'virtual'}
            db_type = type_map.get(self.filters['pallet_type'], self.filters['pallet_type'])
            query += " AND pal.pallet_type = ?"
            params.append(db_type)
        query += " ORDER BY pal.pallet_number, p.package_number, c.component_code"

        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        self.progress_updated.emit(60)

        include_pkg_qr = bool(self.filters.get('include_package_qr'))
        include_pallet_qr = bool(self.filters.get('include_pallet_qr'))

        wb = Workbook()
        ws = wb.active
        ws.title = "托盘-包裹-板件明细"

        base_headers = list(df.columns)
        extra_headers = []
        if include_pallet_qr:
            extra_headers.append("托盘二维码")
        if include_pkg_qr:
            extra_headers.append("包裹二维码")
        headers = base_headers + extra_headers

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        qr_handler = QRCodeHandler()
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
            insert_col = len(base_headers) + 1
            try:
                if include_pallet_qr:
                    pnum = str(row['托盘号'])
                    img = qr_handler.create_qr_code(pnum)
                    bio = BytesIO()
                    img.save(bio, format='PNG')
                    bio.seek(0)
                    xlimg = XLImage(bio)
                    xlimg.width = 80
                    xlimg.height = 80
                    ws.add_image(xlimg, ws.cell(row=row_idx + 2, column=insert_col).coordinate)
                    ws.column_dimensions[ws.cell(row=1, column=insert_col).column_letter].width = 16
                    insert_col += 1
                if include_pkg_qr:
                    pkg = str(row['包裹号'])
                    img = qr_handler.create_qr_code(pkg)
                    bio = BytesIO()
                    img.save(bio, format='PNG')
                    bio.seek(0)
                    xlimg = XLImage(bio)
                    xlimg.width = 80
                    xlimg.height = 80
                    ws.add_image(xlimg, ws.cell(row=row_idx + 2, column=insert_col).coordinate)
                    ws.column_dimensions[ws.cell(row=1, column=insert_col).column_letter].width = 16
                if include_pallet_qr or include_pkg_qr:
                    ws.row_dimensions[row_idx + 2].height = 60
            except Exception:
                pass

        # 列宽调整
        for column in ws.columns:
            try:
                column_letter = column[0].column_letter
                max_length = max(len(str(c.value)) if c.value is not None else 0 for c in column)
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            except Exception:
                pass

        self.progress_updated.emit(90)
        wb.save(self.file_path)
        self.progress_updated.emit(100)

    def export_comprehensive(self):
        """导出综合数据"""
        self.status_updated.emit("正在查询综合数据...")
        
        conn = db.get_connection()
        
        # 查询所有板件的完整追踪信息
        query = '''
            SELECT 
                c.component_code,
                c.component_name,
                c.material,
                c.finished_size,
                c.room_number,
                c.cabinet_number,
                o.order_number,
                o.customer_name,
                o.customer_address,
                p.package_number,
                p.status as package_status,
                p.created_at as package_created_at,
                pal.pallet_number,
                pal.pallet_type,
                pal.status as pallet_status,
                c.scanned_at,
                c.status as component_status
            FROM components c
            LEFT JOIN orders o ON c.order_id = o.id
            LEFT JOIN packages p ON c.package_id = p.id
            LEFT JOIN pallet_packages pp ON pp.package_id = p.id
            LEFT JOIN pallets pal ON pp.pallet_id = pal.id
            WHERE 1=1
        '''
        
        params = []
        
        # 添加过滤条件
        if self.filters.get('start_date'):
            query += " AND DATE(c.created_at) >= ?"
            params.append(self.filters['start_date'])
        
        if self.filters.get('end_date'):
            query += " AND DATE(c.created_at) <= ?"
            params.append(self.filters['end_date'])
        
        if self.filters.get('order_number'):
            query += " AND o.order_number LIKE ?"
            params.append(f"%{self.filters['order_number']}%")
        
        query += " ORDER BY c.created_at DESC"
        
        df_comprehensive = pd.read_sql_query(query, conn, params=params)
        conn.close()
        
        self.progress_updated.emit(50)
        
        # 创建Excel文件
        self.status_updated.emit("正在生成Excel文件...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "板件追踪信息"
        
        # 设置表头
        headers = ['板件编码', '板件名称', '材质', '尺寸', '房间号', '柜号',
                  '订单号', '客户名称', '客户地址', '包装号', '包装状态', '包装时间',
                  '托盘号', '托盘类型', '托盘状态', '扫描时间', '板件状态']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        self.progress_updated.emit(70)
        
        # 填充数据
        for row_idx, row in df_comprehensive.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx + 2, column=col_idx, value=value)
        
        self.progress_updated.emit(90)
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存文件
        wb.save(self.file_path)
        self.progress_updated.emit(100)

class ExportDialog(QDialog):
    """导出对话框"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("数据导出")
        self.setModal(True)
        self.resize(500, 400)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 导出类型
        type_group = QGroupBox("导出类型")
        type_layout = QVBoxLayout(type_group)
        
        self.pallets_radio = QCheckBox("托盘数据")
        type_layout.addWidget(self.pallets_radio)

        self.package_components_radio = QCheckBox("包裹-板件明细")
        type_layout.addWidget(self.package_components_radio)
        
        self.pallet_package_components_radio = QCheckBox("托盘-包裹-板件明细")
        type_layout.addWidget(self.pallet_package_components_radio)
        
        layout.addWidget(type_group)
        
        # 二维码选项
        qr_group = QGroupBox("二维码选项")
        qr_layout = QGridLayout(qr_group)
        self.include_package_qr_cb = QCheckBox("导出包裹二维码")
        self.include_package_qr_cb.setChecked(False)
        qr_layout.addWidget(self.include_package_qr_cb, 0, 0)
        self.include_pallet_qr_cb = QCheckBox("导出托盘二维码")
        self.include_pallet_qr_cb.setChecked(False)
        qr_layout.addWidget(self.include_pallet_qr_cb, 0, 1)
        layout.addWidget(qr_group)
        
        # 文件路径
        path_group = QGroupBox("保存路径")
        path_layout = QHBoxLayout(path_group)
        
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("选择保存路径...")
        path_layout.addWidget(self.file_path_edit)
        
        self.browse_btn = QPushButton("浏览...")
        self.browse_btn.clicked.connect(self.browse_file)
        path_layout.addWidget(self.browse_btn)
        
        layout.addWidget(path_group)
        
        # 按钮
        button_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("开始导出")
        self.export_btn.clicked.connect(self.start_export)
        button_layout.addWidget(self.export_btn)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
    
    def browse_file(self):
        """浏览文件"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "", "Excel文件 (*.xlsx)")
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            self.file_path_edit.setText(file_path)
    
    def start_export(self):
        """开始导出"""
        # 检查导出类型
        export_types = []
        if self.pallets_radio.isChecked():
            export_types.append("pallets")
        if self.package_components_radio.isChecked():
            export_types.append("package_components")
        if self.pallet_package_components_radio.isChecked():
            export_types.append("pallet_package_components")
        
        if not export_types:
            QMessageBox.warning(self, "警告", "请选择至少一种导出类型")
            return
        
        # 检查文件路径
        file_path = self.file_path_edit.text().strip()
        if not file_path:
            QMessageBox.warning(self, "警告", "请选择保存路径")
            return
        
        # 准备过滤条件（简化版，不包含日期和订单过滤）
        filters = {
            'include_package_qr': self.include_package_qr_cb.isChecked(),
            'include_pallet_qr': self.include_pallet_qr_cb.isChecked(),
        }
        
        # 开始导出
        self.export_data(export_types, file_path, filters)
    
    def export_data(self, export_types, base_file_path, filters):
        """导出数据"""
        for i, export_type in enumerate(export_types):
            # 生成文件名
            if len(export_types) > 1:
                name_map = {
                    "pallets": "托盘数据", 
                    "package_components": "包裹-板件明细",
                    "pallet_package_components": "托盘-包裹-板件明细"
                }
                file_path = base_file_path.replace('.xlsx', f'_{name_map[export_type]}.xlsx')
            else:
                file_path = base_file_path
            
            # 创建进度对话框
            progress = QProgressDialog(f"正在导出{export_type}数据...", "取消", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)
            
            # 创建工作线程
            worker = ExportWorker(export_type, file_path, filters)
            worker.progress_updated.connect(progress.setValue)
            worker.status_updated.connect(progress.setLabelText)
            worker.export_finished.connect(lambda success, msg: self.on_export_finished(success, msg, progress))
            
            # 启动导出
            worker.start()
            
            # 显示进度对话框
            if progress.exec_() == QProgressDialog.Rejected:
                worker.terminate()
                return
            
            worker.wait()
    
    def on_export_finished(self, success, message, progress):
        """导出完成"""
        progress.close()
        
        if success:
            QMessageBox.information(self, "成功", message)
            self.accept()
        else:
            QMessageBox.critical(self, "错误", message)

class BarChartWidget(QWidget):
    """简易水平柱状图控件"""
    def __init__(self, title="", parent=None):
        super().__init__(parent)
        self.title = title
        self.data = []  # [(label, value)]
        self.setMinimumHeight(180)
        self.palette = [
            QColor(33, 150, 243),  # 蓝
            QColor(76, 175, 80),   # 绿
            QColor(255, 152, 0),   # 橙
            QColor(156, 39, 176),  # 紫
            QColor(244, 67, 54),   # 红
            QColor(0, 188, 212),   # 青
        ]

    def set_data(self, items):
        """设置数据: items为[(label, value)]"""
        self.data = items or []
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        rect = self.rect()

        margin = 16
        title_h = 22
        bar_h = 22
        spacing = 10

        # 标题
        painter.setPen(Qt.black)
        title_font = QFont()
        title_font.setPointSize(10)
        title_font.setBold(True)
        painter.setFont(title_font)
        painter.drawText(rect.adjusted(margin, 4, -margin, -4), Qt.AlignLeft | Qt.AlignTop, self.title)

        max_val = max([v for _, v in self.data], default=0)
        if max_val == 0:
            painter.drawText(rect, Qt.AlignCenter, "无数据")
            return

        # 条形区域宽度（预留标签空间）
        label_w = 100
        bar_area_w = max(0, rect.width() - 2 * margin - label_w - 40)

        painter.setFont(QFont("", 9))
        total = sum([v for _, v in self.data])
        y = margin + title_h
        for i, (label, value) in enumerate(self.data):
            # 标签
            painter.drawText(margin, y + bar_h - 6, str(label))

            # 条形
            bw = int(bar_area_w * (value / max_val))
            bar_x = margin + label_w
            color = self.palette[i % len(self.palette)]
            painter.fillRect(bar_x, y, bw, bar_h, color)
            painter.drawRect(bar_x, y, bw, bar_h)

            # 数值
            percent = (value / total * 100) if total else 0
            painter.drawText(bar_x + bw + 6, y + bar_h - 6, f"{value} ({percent:.1f}%)")
            y += bar_h + spacing

        # 颜色图例
        legend_y = y + 6
        legend_x = margin
        painter.setFont(QFont("", 8))
        for i, (label, _) in enumerate(self.data):
            color = self.palette[i % len(self.palette)]
            painter.fillRect(legend_x, legend_y, 12, 12, color)
            painter.drawRect(legend_x, legend_y, 12, 12)
            painter.drawText(legend_x + 18, legend_y + 11, str(label))
            legend_x += 120

class Reports(QWidget):
    """报表统计模块"""
    
    def __init__(self):
        super().__init__()
        self.selected_order_id = None
        self._selected_order_text = None
        self.init_ui()
        # 直接加载统计，无需下拉框
        self.load_statistics()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 工具栏
        toolbar_layout = QHBoxLayout()
        
        self.export_btn = QPushButton("数据导出")
        self.export_btn.clicked.connect(self.export_data)
        toolbar_layout.addWidget(self.export_btn)
        
        self.refresh_btn = QPushButton("刷新统计")
        self.refresh_btn.clicked.connect(self.load_statistics)
        toolbar_layout.addWidget(self.refresh_btn)

        # 简化的订单选择按钮（直接显示当前选择的订单号）
        self.select_order_btn = QPushButton("选择订单…")
        self.select_order_btn.setToolTip("打开订单选择对话框")
        self.select_order_btn.clicked.connect(self.open_order_selection_dialog)
        toolbar_layout.addWidget(self.select_order_btn)
        # 移除旧的下拉搜索与容器，仅保留按钮
        
        # 在工具栏右侧显示当前订单标签
        toolbar_layout.addStretch()
        self.current_order_label = QLabel("当前订单：全部订单")
        self.current_order_label.setToolTip("当前报表筛选的订单")
        toolbar_layout.addWidget(self.current_order_label)
        
        layout.addLayout(toolbar_layout)
        
        # 统计信息标签页
        self.tab_widget = QTabWidget()
        layout.addWidget(self.tab_widget)
        
        # 概览统计
        self.init_overview_tab()
        
        # 订单统计
        self.init_orders_tab()
        
        # 包装统计
        self.init_packages_tab()
        
        # 托盘统计
        self.init_pallets_tab()
    
    def init_overview_tab(self):
        """概览统计标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # 顶部统计卡片区域
        stats_frame = QGroupBox("数据概览")
        stats_layout = QGridLayout(stats_frame)
        
        # 创建统计卡片
        self.create_stat_card(stats_layout, "总订单数", "total_orders", "#2196F3", 0, 0)
        self.create_stat_card(stats_layout, "总板件数", "total_components", "#4CAF50", 0, 1)
        self.create_stat_card(stats_layout, "总包装数", "total_packages", "#FF9800", 0, 2)
        self.create_stat_card(stats_layout, "总托盘数", "total_pallets", "#9C27B0", 0, 3)
        
        # 第二行统计卡片
        self.create_stat_card(stats_layout, "已打包板件", "packaged_components", "#00BCD4", 1, 0)
        self.create_stat_card(stats_layout, "未打包板件", "unpackaged_components", "#F44336", 1, 1)
        self.create_stat_card(stats_layout, "已封装包装", "sealed_packages", "#8BC34A", 1, 2)
        self.create_stat_card(stats_layout, "已封托托盘", "sealed_pallets", "#673AB7", 1, 3)
        
        layout.addWidget(stats_frame)
        
        # 图表展示区域
        charts_frame = QGroupBox("数据分析")
        charts_layout = QHBoxLayout(charts_frame)
        
        # 包装状态分布图
        self.packages_chart = BarChartWidget("包装状态分布")
        charts_layout.addWidget(self.packages_chart)
        
        # 托盘类型分布图
        self.pallets_chart = BarChartWidget("托盘类型分布")
        charts_layout.addWidget(self.pallets_chart)
        
        # 订单进度图
        self.orders_progress_chart = BarChartWidget("订单完成进度")
        charts_layout.addWidget(self.orders_progress_chart)
        
        layout.addWidget(charts_frame)
        
        # 添加刷新按钮
        refresh_layout = QHBoxLayout()
        refresh_layout.addStretch()
        self.refresh_btn = QPushButton("刷新数据")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        self.refresh_btn.clicked.connect(self.load_statistics)
        refresh_layout.addWidget(self.refresh_btn)
        layout.addLayout(refresh_layout)
        
        self.tab_widget.addTab(tab, "概览统计")
    
    def create_stat_card(self, layout, title, attr_name, color, row, col):
        """创建统计卡片"""
        # 创建容器
        card_frame = QGroupBox()
        card_frame.setStyleSheet(f"""
            QGroupBox {{
                border: 2px solid {color};
                border-radius: 8px;
                margin: 5px;
                padding: 10px;
                background-color: white;
            }}
            QGroupBox:hover {{
                background-color: #f5f5f5;
                border-color: {color};
            }}
        """)
        
        card_layout = QVBoxLayout(card_frame)
        card_layout.setSpacing(5)
        
        # 标题
        title_label = QLabel(title)
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(f"""
            font-size: 14px; 
            font-weight: bold; 
            color: {color};
            margin-bottom: 5px;
        """)
        
        # 数值
        value_label = QLabel("0")
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setStyleSheet(f"""
            font-size: 28px; 
            font-weight: bold; 
            color: {color};
            margin: 10px 0;
        """)
        
        # 设置属性名以便后续更新
        setattr(self, f"{attr_name}_label", value_label)
        
        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)
        
        layout.addWidget(card_frame, row, col)

    def init_orders_tab(self):
        """订单统计标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(6)
        self.orders_table.setHorizontalHeaderLabels([
            '订单号', '客户名称', '板件数量', '已打包', '未打包', '创建时间'
        ])
        self.orders_table.horizontalHeader().setStretchLastSection(True)
        self.orders_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.orders_table.setAlternatingRowColors(True)
        
        # 设置滚动条策略
        self.orders_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.orders_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        
        layout.addWidget(self.orders_table)
        self.tab_widget.addTab(tab, "订单统计")
    
    def init_packages_tab(self):
        """包装统计标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        self.packages_table = QTableWidget()
        self.packages_table.setColumnCount(6)
        self.packages_table.setHorizontalHeaderLabels([
            '包装号', '订单号', '客户名称', '板件数量', '状态', '创建时间'
        ])
        self.packages_table.horizontalHeader().setStretchLastSection(True)
        self.packages_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.packages_table.setAlternatingRowColors(True)
        
        layout.addWidget(self.packages_table)
        self.tab_widget.addTab(tab, "包装统计")
    
    def init_pallets_tab(self):
        """托盘统计标签页"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        self.pallets_table = QTableWidget()
        self.pallets_table.setColumnCount(6)
        self.pallets_table.setHorizontalHeaderLabels([
            '托盘号', '托盘类型', '包装数量', '板件数量', '状态', '创建时间'
        ])
        self.pallets_table.horizontalHeader().setStretchLastSection(True)
        self.pallets_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.pallets_table.setAlternatingRowColors(True)
        
        layout.addWidget(self.pallets_table)
        self.tab_widget.addTab(tab, "托盘统计")

    # 已移除下拉框与搜索逻辑，改为仅按钮选择订单

    def open_order_selection_dialog(self):
        """打开订单选择对话框并回填选择"""
        try:
            dialog = OrderSelectionDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                selected = dialog.get_selected_order()
                if selected:
                    # 设置选中订单ID
                    try:
                        self.selected_order_id = int(selected.get('id')) if selected.get('id') is not None else None
                    except Exception:
                        self.selected_order_id = selected.get('id')
                    # 记录当前订单文本
                    self._selected_order_text = f"{selected.get('order_number','')} - {selected.get('customer_name','')}".strip()
                    # 直接把订单号显示在按钮上
                    try:
                        self.select_order_btn.setText(f"订单: {selected.get('order_number','')}")
                    except Exception:
                        pass
                    # 更新右侧标签并刷新统计
                    self.update_current_order_label()
                    self.load_statistics()
        except Exception as e:
            QMessageBox.warning(self, "警告", f"打开订单选择失败：{str(e)}")

    def update_current_order_label(self):
        """更新工具栏右侧的当前订单只读标签"""
        try:
            if self.selected_order_id:
                # 优先使用对话框选择记录的文本
                text = getattr(self, '_selected_order_text', None)
                order_number = text.split(' - ')[0] if text else ''
                self.current_order_label.setText(f"当前订单：{order_number or '未知'}")
            else:
                self.current_order_label.setText("当前订单：全部订单")
        except Exception:
            # 标签更新失败时保持不影响主流程
            pass
    # 删除搜索补全相关逻辑

    # 状态与类型中文映射
    def _map_package_status(self, status):
        # 统一映射，兼容旧别名 packed→sealed
        try:
            from status_utils import package_status_cn
            return package_status_cn(status)
        except Exception:
            mapping = {
                'completed': '已完成',
                'open': '进行中',
                'sealed': '已封包',
                'packed': '已封包',
            }
            if status is None or str(status).strip() == '':
                return '未设置'
            return mapping.get(str(status), str(status))

    def _map_pallet_status(self, status):
        try:
            from status_utils import pallet_status_cn
            return pallet_status_cn(status)
        except Exception:
            mapping = {
                'open': '开放',
                'sealed': '已封托',
                'closed': '已关闭',
            }
            if status is None or str(status).strip() == '':
                return '未设置'
            return mapping.get(str(status), str(status))

    def _map_pallet_type(self, ptype):
        if ptype == 'virtual' or (isinstance(ptype, str) and ptype.startswith('VT')):
            return '虚拟托盘'
        if ptype == 'physical' or (isinstance(ptype, str) and not ptype.startswith('VT')):
            return '实体托盘'
        return str(ptype)
    
    def load_statistics(self):
        """加载统计数据"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 概览统计
        if self.selected_order_id:
            cursor.execute("SELECT COUNT(*) FROM orders WHERE id = ?", (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM orders")
        total_orders = cursor.fetchone()[0]
        self.total_orders_label.setText(str(total_orders))
        
        if self.selected_order_id:
            cursor.execute("SELECT COUNT(*) FROM components WHERE order_id = ?", (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM components")
        total_components = cursor.fetchone()[0]
        self.total_components_label.setText(str(total_components))
        
        if self.selected_order_id:
            cursor.execute("SELECT COUNT(*) FROM packages WHERE order_id = ?", (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM packages")
        total_packages = cursor.fetchone()[0]
        self.total_packages_label.setText(str(total_packages))
        
        if self.selected_order_id:
            cursor.execute('''
                SELECT COUNT(DISTINCT pal.id)
                FROM pallets pal
                LEFT JOIN pallet_packages pp ON pp.pallet_id = pal.id
                LEFT JOIN packages p ON pp.package_id = p.id
                WHERE p.order_id = ?
            ''', (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM pallets")
        total_pallets = cursor.fetchone()[0]
        self.total_pallets_label.setText(str(total_pallets))
        
        # 新增统计卡片数据
        # 已打包板件
        if self.selected_order_id:
            cursor.execute("SELECT COUNT(*) FROM components WHERE order_id = ? AND package_id IS NOT NULL", (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM components WHERE package_id IS NOT NULL")
        packaged_components = cursor.fetchone()[0]
        self.packaged_components_label.setText(str(packaged_components))
        
        # 未打包板件
        if self.selected_order_id:
            cursor.execute("SELECT COUNT(*) FROM components WHERE order_id = ? AND package_id IS NULL", (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM components WHERE package_id IS NULL")
        unpackaged_components = cursor.fetchone()[0]
        self.unpackaged_components_label.setText(str(unpackaged_components))
        
        # 已封装包装
        if self.selected_order_id:
            cursor.execute("SELECT COUNT(*) FROM packages WHERE order_id = ? AND status = 'sealed'", (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM packages WHERE status = 'sealed'")
        sealed_packages = cursor.fetchone()[0]
        self.sealed_packages_label.setText(str(sealed_packages))
        
        # 已封托托盘
        if self.selected_order_id:
            cursor.execute('''
                SELECT COUNT(DISTINCT pal.id)
                FROM pallets pal
                LEFT JOIN pallet_packages pp ON pp.pallet_id = pal.id
                LEFT JOIN packages p ON pp.package_id = p.id
                WHERE p.order_id = ? AND pal.status = 'sealed'
            ''', (self.selected_order_id,))
        else:
            cursor.execute("SELECT COUNT(*) FROM pallets WHERE status = 'sealed'")
        sealed_pallets = cursor.fetchone()[0]
        self.sealed_pallets_label.setText(str(sealed_pallets))
        
        # 图表数据
        # 包装状态分布
        if self.selected_order_id:
            cursor.execute('''
                SELECT status, COUNT(*) 
                FROM packages 
                WHERE order_id = ? 
                GROUP BY status
            ''', (self.selected_order_id,))
        else:
            cursor.execute("SELECT status, COUNT(*) FROM packages GROUP BY status")
        package_status_data = [(self._map_package_status(status), count) for status, count in cursor.fetchall()]
        self.packages_chart.set_data(package_status_data)
        
        # 托盘类型分布
        if self.selected_order_id:
            cursor.execute('''
                SELECT pal.pallet_type, COUNT(DISTINCT pal.id)
                FROM pallets pal
                LEFT JOIN pallet_packages pp ON pp.pallet_id = pal.id
                LEFT JOIN packages p ON pp.package_id = p.id
                WHERE p.order_id = ?
                GROUP BY pal.pallet_type
            ''', (self.selected_order_id,))
        else:
            cursor.execute("SELECT pallet_type, COUNT(*) FROM pallets GROUP BY pallet_type")
        pallet_type_data = [(self._map_pallet_type(ptype), count) for ptype, count in cursor.fetchall()]
        self.pallets_chart.set_data(pallet_type_data)
        
        # 订单完成进度
        if self.selected_order_id:
            cursor.execute('''
                SELECT 
                    o.order_number,
                    ROUND(CAST(COUNT(CASE WHEN c.package_id IS NOT NULL THEN 1 END) AS FLOAT) / COUNT(c.id) * 100, 1) as progress
                FROM orders o
                LEFT JOIN components c ON c.order_id = o.id
                WHERE o.id = ?
                GROUP BY o.id
            ''', (self.selected_order_id,))
        else:
            cursor.execute('''
                SELECT 
                    o.order_number,
                    ROUND(CAST(COUNT(CASE WHEN c.package_id IS NOT NULL THEN 1 END) AS FLOAT) / COUNT(c.id) * 100, 1) as progress
                FROM orders o
                LEFT JOIN components c ON c.order_id = o.id
                GROUP BY o.id
                ORDER BY progress DESC
                LIMIT 10
            ''')
        order_progress_data = [(order_num, progress) for order_num, progress in cursor.fetchall() if progress is not None]
        self.orders_progress_chart.set_data(order_progress_data)
        
        # 订单统计
        orders_query = '''
            SELECT 
                o.order_number,
                o.customer_name,
                COUNT(c.id) as total_components,
                COUNT(CASE WHEN c.package_id IS NOT NULL THEN 1 END) as packaged,
                COUNT(CASE WHEN c.package_id IS NULL THEN 1 END) as unpackaged,
                o.created_at
            FROM orders o
            LEFT JOIN components c ON c.order_id = o.id
        '''
        params = []
        if self.selected_order_id:
            orders_query += " WHERE o.id = ?"
            params.append(self.selected_order_id)
        orders_query += " GROUP BY o.id ORDER BY o.created_at DESC"
        cursor.execute(orders_query, params)
        orders_data = cursor.fetchall()
        
        self.orders_table.setRowCount(len(orders_data))
        for i, row in enumerate(orders_data):
            for j, value in enumerate(row):
                self.orders_table.setItem(i, j, QTableWidgetItem(str(value) if value else ''))
        
        # 包装统计
        packages_query = '''
            SELECT 
                p.package_number,
                o.order_number,
                o.customer_name,
                COUNT(c.id) as component_count,
                p.status,
                p.created_at
            FROM packages p
            LEFT JOIN orders o ON p.order_id = o.id
            LEFT JOIN components c ON c.package_id = p.id
        '''
        params = []
        if self.selected_order_id:
            packages_query += " WHERE p.order_id = ?"
            params.append(self.selected_order_id)
        packages_query += " GROUP BY p.id ORDER BY p.created_at DESC LIMIT 100"
        cursor.execute(packages_query, params)
        packages_data = cursor.fetchall()
        
        self.packages_table.setRowCount(len(packages_data))
        for i, row in enumerate(packages_data):
            for j, value in enumerate(row):
                text = str(value) if value else ''
                if j == 4:  # 状态列
                    text = self._map_package_status(value)
                self.packages_table.setItem(i, j, QTableWidgetItem(text))
        
        # 托盘统计
        pallets_query = '''
            SELECT 
                pal.pallet_number,
                pal.pallet_type,
                COUNT(DISTINCT p.id) as package_count,
                COUNT(c.id) as component_count,
                pal.status,
                pal.created_at
            FROM pallets pal
            LEFT JOIN packages p ON p.pallet_id = pal.id
            LEFT JOIN components c ON c.package_id = p.id
        '''
        params = []
        if self.selected_order_id:
            pallets_query += " WHERE p.order_id = ?"
            params.append(self.selected_order_id)
        pallets_query += " GROUP BY pal.id ORDER BY pal.created_at DESC LIMIT 100"
        cursor.execute(pallets_query, params)
        pallets_data = cursor.fetchall()
        
        self.pallets_table.setRowCount(len(pallets_data))
        for i, row in enumerate(pallets_data):
            for j, value in enumerate(row):
                text = str(value) if value else ''
                if j == 1:  # 托盘类型
                    text = self._map_pallet_type(value)
                elif j == 4:  # 托盘状态
                    text = self._map_pallet_status(value)
                self.pallets_table.setItem(i, j, QTableWidgetItem(text))
        
        conn.close()
    
    def export_data(self):
        """导出数据"""
        dialog = ExportDialog(self)
        # 与报表页订单筛选联动：预填订单号
        try:
            if self.selected_order_id:
                text = getattr(self, '_selected_order_text', None)
                if not text:
                    # 兜底：从数据库查询订单号
                    conn = db.get_connection()
                    cur = conn.cursor()
                    cur.execute('SELECT order_number FROM orders WHERE id = ?', (self.selected_order_id,))
                    row = cur.fetchone()
                    conn.close()
                    text = (row[0] if row else '')
                order_number = text.split(' - ')[0] if isinstance(text, str) and ' - ' in text else str(text or '')
                dialog.order_number_edit.setText(order_number)
        except Exception:
            pass
        # 同步联动托盘类型：基于当前统计结果预选
        try:
            # 从概览图表获取托盘类型分布，选择占比最高的类型作为默认值
            items = getattr(self.pallets_chart, 'data', [])
            if items:
                # items: [(label, value)]，label为中文：'实体托盘' 或 '虚拟托盘'
                top_label = max(items, key=lambda x: x[1])[0]
                # 设置下拉框选中项
                index = dialog.pallet_type_combo.findText(str(top_label))
                if index >= 0:
                    dialog.pallet_type_combo.setCurrentIndex(index)
        except Exception:
            # 忽略联动异常，确保导出对话框正常打开
            pass
        dialog.exec_()