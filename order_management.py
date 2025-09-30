import sys
import csv
import json
from datetime import datetime
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
                             QPushButton, QTableWidget, QTableWidgetItem, QLabel,
                             QLineEdit, QTextEdit, QComboBox, QFileDialog, QMessageBox,
                             QDialog, QDialogButtonBox, QGroupBox, QCheckBox,
                             QProgressBar, QSplitter, QHeaderView, QTabWidget, QFormLayout,
                             QInputDialog, QScrollArea, QAbstractItemView)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QEvent
from PyQt5.QtGui import QFont
from database import db

class OrderSelectionDialog(QDialog):
    """订单选择对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择订单")
        self.setModal(True)
        self.resize(600, 400)
        self.selected_order = None
        
        self.init_ui()
        self.load_orders()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 说明文本
        info_label = QLabel("请选择要导入数据的订单：")
        info_label.setStyleSheet("font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(info_label)
        
        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入订单号或客户名称进行搜索...")
        self.search_edit.textChanged.connect(self.filter_orders)
        search_layout.addWidget(self.search_edit)
        layout.addLayout(search_layout)
        
        # 订单表格
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(4)
        self.orders_table.setHorizontalHeaderLabels(['订单号', '客户名称', '客户地址', '创建时间'])
        self.orders_table.horizontalHeader().setStretchLastSection(True)
        self.orders_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.orders_table.setSelectionMode(QTableWidget.SingleSelection)
        # 禁止编辑，避免单击进入编辑模式导致选择失效
        self.orders_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.orders_table.setAlternatingRowColors(True)
        self.orders_table.itemDoubleClicked.connect(self.on_order_double_clicked)
        # 单击行立即选中并启用“选择”按钮
        self.orders_table.itemClicked.connect(self.on_order_clicked)
        
        # 设置滚动条策略
        self.orders_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.orders_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        layout.addWidget(self.orders_table)
        
        # 按钮
        button_layout = QHBoxLayout()
        
        self.new_order_btn = QPushButton("新建订单")
        self.new_order_btn.clicked.connect(self.new_order)
        button_layout.addWidget(self.new_order_btn)
        
        button_layout.addStretch()
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        self.select_btn = QPushButton("选择")
        self.select_btn.clicked.connect(self.select_order)
        self.select_btn.setEnabled(False)
        # 设为默认按钮，支持回车直接确认
        self.select_btn.setAutoDefault(True)
        self.select_btn.setDefault(True)
        button_layout.addWidget(self.select_btn)
        
        layout.addLayout(button_layout)
        
        # 连接选择变化事件
        self.orders_table.itemSelectionChanged.connect(self.on_selection_changed)
    
    def load_orders(self):
        """加载订单列表"""
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, order_number, customer_name, customer_address, created_at
                FROM orders
                ORDER BY created_at DESC
            ''')
            orders = cursor.fetchall()
            conn.close()
            
            # 存储原始数据用于搜索
            self.all_orders = orders
            self.display_orders(orders)
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载订单列表失败：\n{str(e)}")
    
    def display_orders(self, orders):
        """显示订单列表"""
        self.orders_table.setRowCount(len(orders))
        for i, order in enumerate(orders):
            self.orders_table.setItem(i, 0, QTableWidgetItem(order[1]))  # order_number
            self.orders_table.setItem(i, 1, QTableWidgetItem(order[2] or ''))  # customer_name
            self.orders_table.setItem(i, 2, QTableWidgetItem(order[3] or ''))  # customer_address
            self.orders_table.setItem(i, 3, QTableWidgetItem(order[4]))  # created_at
            
            # 存储订单ID
            self.orders_table.item(i, 0).setData(Qt.UserRole, order[0])
        # 默认选中第一行，减少多次点击
        if orders:
            self.orders_table.selectRow(0)
            self.orders_table.setCurrentCell(0, 0)
            self.select_btn.setEnabled(True)
            self.orders_table.setFocus()
    
    def filter_orders(self):
        """根据搜索文本过滤订单"""
        search_text = self.search_edit.text().lower()
        if not search_text:
            # 如果搜索框为空，显示所有订单
            self.display_orders(self.all_orders)
        else:
            # 过滤订单
            filtered_orders = []
            for order in self.all_orders:
                order_number = order[1].lower() if order[1] else ''
                customer_name = order[2].lower() if order[2] else ''
                if search_text in order_number or search_text in customer_name:
                    filtered_orders.append(order)
            self.display_orders(filtered_orders)
    
    def on_selection_changed(self):
        """选择变化时启用/禁用选择按钮"""
        self.select_btn.setEnabled(self.orders_table.currentRow() >= 0)

    def on_order_double_clicked(self, item):
        """双击订单时直接选择"""
        self.select_order()
    
    def on_order_clicked(self, item):
        """单击行时确保选中并启用选择按钮"""
        try:
            if item is not None:
                row = item.row()
                if row >= 0:
                    self.orders_table.selectRow(row)
                    self.select_btn.setEnabled(True)
        except Exception:
            pass
    
    def select_order(self):
        """选择订单"""
        current_row = self.orders_table.currentRow()
        if current_row >= 0:
            order_id = self.orders_table.item(current_row, 0).data(Qt.UserRole)
            order_number = self.orders_table.item(current_row, 0).text()
            customer_name = self.orders_table.item(current_row, 1).text()
            customer_address = self.orders_table.item(current_row, 2).text()
            
            self.selected_order = {
                'id': order_id,
                'order_number': order_number,
                'customer_name': customer_name,
                'customer_address': customer_address
            }
            self.accept()
    
    def new_order(self):
        """新建订单"""
        dialog = NewOrderDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            order_data = dialog.get_order_data()
            order_id = order_data['id']
            
            # 重新加载订单列表
            self.load_orders()
            
            # 自动选择新创建的订单
            for i in range(self.orders_table.rowCount()):
                if self.orders_table.item(i, 0).data(Qt.UserRole) == order_id:
                    self.orders_table.selectRow(i)
                    break
            
            QMessageBox.information(self, "成功", "订单创建成功")
    
    def get_selected_order(self):
        """获取选中的订单"""
        return self.selected_order

class EditOrderDialog(QDialog):
    """编辑订单对话框"""
    def __init__(self, parent=None, order_id=None):
        super().__init__(parent)
        self.order_id = order_id
        self.setWindowTitle("编辑订单")
        self.setModal(True)
        self.resize(400, 300)
        
        self.init_ui()
        self.load_order_data()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # 订单号
        self.order_number_edit = QLineEdit()
        form_layout.addRow("订单号:", self.order_number_edit)
        
        # 客户名称
        self.customer_name_edit = QLineEdit()
        form_layout.addRow("客户名称:", self.customer_name_edit)
        
        # 客户地址
        self.customer_address_edit = QTextEdit()
        self.customer_address_edit.setMaximumHeight(80)
        form_layout.addRow("客户地址:", self.customer_address_edit)
        
        layout.addLayout(form_layout)
        
        # 按钮
        button_layout = QHBoxLayout()
        
        self.save_btn = QPushButton("保存")
        self.save_btn.clicked.connect(self.save_order)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.save_btn)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
    
    def load_order_data(self):
        """加载订单数据"""
        if not self.order_id:
            return
        
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT order_number, customer_name, customer_address
                FROM orders WHERE id = ?
            ''', (self.order_id,))
            
            order = cursor.fetchone()
            conn.close()
            
            if order:
                self.order_number_edit.setText(order[0] or '')
                self.customer_name_edit.setText(order[1] or '')
                self.customer_address_edit.setPlainText(order[2] or '')
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载订单数据失败：\n{str(e)}")
    
    def save_order(self):
        """保存订单"""
        order_number = self.order_number_edit.text().strip()
        customer_name = self.customer_name_edit.text().strip()
        customer_address = self.customer_address_edit.toPlainText().strip()
        
        if not order_number:
            QMessageBox.warning(self, "警告", "请输入订单号")
            return
        
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # 检查订单号是否重复（排除当前订单）
            cursor.execute('''
                SELECT id FROM orders WHERE order_number = ? AND id != ?
            ''', (order_number, self.order_id))
            
            if cursor.fetchone():
                QMessageBox.warning(self, "警告", "订单号已存在，请使用其他订单号")
                conn.close()
                return
            
            # 更新订单
            cursor.execute('''
                UPDATE orders 
                SET order_number = ?, customer_name = ?, customer_address = ?
                WHERE id = ?
            ''', (order_number, customer_name, customer_address, self.order_id))
            
            conn.commit()
            conn.close()
            
            # 记录操作日志
            db.log_operation('edit_order', {
                'order_id': self.order_id,
                'order_number': order_number,
                'customer_name': customer_name,
                'customer_address': customer_address
            })
            
            QMessageBox.information(self, "成功", "订单信息已更新")
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存订单失败：\n{str(e)}")

class NewOrderDialog(QDialog):
    """新建订单对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("新建订单")
        self.setModal(True)
        self.resize(400, 300)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 订单信息组
        order_group = QGroupBox("订单信息")
        order_layout = QFormLayout(order_group)
        
        # 订单号
        self.order_number_edit = QLineEdit()
        self.order_number_edit.setPlaceholderText("请输入订单号")
        order_layout.addRow("订单号*:", self.order_number_edit)
        
        # 客户名称
        self.customer_name_edit = QLineEdit()
        self.customer_name_edit.setPlaceholderText("请输入客户名称")
        order_layout.addRow("客户名称:", self.customer_name_edit)
        
        # 客户地址
        self.customer_address_edit = QTextEdit()
        self.customer_address_edit.setPlaceholderText("请输入客户地址")
        self.customer_address_edit.setMaximumHeight(80)
        order_layout.addRow("客户地址:", self.customer_address_edit)
        
        # 联系电话
        self.customer_phone_edit = QLineEdit()
        self.customer_phone_edit.setPlaceholderText("请输入联系电话")
        order_layout.addRow("联系电话:", self.customer_phone_edit)
        
        # 备注
        self.notes_edit = QTextEdit()
        self.notes_edit.setPlaceholderText("请输入备注信息")
        self.notes_edit.setMaximumHeight(60)
        order_layout.addRow("备注:", self.notes_edit)
        
        layout.addWidget(order_group)
        
        # 按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        self.create_btn = QPushButton("创建订单")
        self.create_btn.clicked.connect(self.create_order)
        self.create_btn.setDefault(True)
        button_layout.addWidget(self.create_btn)
        
        layout.addLayout(button_layout)
        
        # 设置焦点
        self.order_number_edit.setFocus()
    
    def create_order(self):
        """创建订单"""
        order_number = self.order_number_edit.text().strip()
        if not order_number:
            QMessageBox.warning(self, "警告", "请输入订单号")
            self.order_number_edit.setFocus()
            return
        
        customer_name = self.customer_name_edit.text().strip()
        customer_address = self.customer_address_edit.toPlainText().strip()
        customer_phone = self.customer_phone_edit.text().strip()
        notes = self.notes_edit.toPlainText().strip()
        
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # 检查订单号是否已存在
            cursor.execute('SELECT id FROM orders WHERE order_number = ?', (order_number,))
            if cursor.fetchone():
                QMessageBox.warning(self, "警告", "订单号已存在，请使用其他订单号")
                self.order_number_edit.setFocus()
                return
            
            # 创建订单
            cursor.execute('''
                INSERT INTO orders (order_number, customer_name, customer_address, customer_phone, notes)
                VALUES (?, ?, ?, ?, ?)
            ''', (order_number, customer_name, customer_address, customer_phone, notes))
            
            self.order_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"创建订单失败：\n{str(e)}")
    
    def get_order_data(self):
        """获取订单数据"""
        return {
            'id': getattr(self, 'order_id', None),
            'order_number': self.order_number_edit.text().strip(),
            'customer_name': self.customer_name_edit.text().strip(),
            'customer_address': self.customer_address_edit.toPlainText().strip(),
            'customer_phone': self.customer_phone_edit.text().strip(),
            'notes': self.notes_edit.toPlainText().strip()
        }

class ImportConfigDialog(QDialog):
    """清单导入配置对话框"""
    def __init__(self, parent=None, csv_file_path=None):
        super().__init__(parent)
        self.setWindowTitle("清单导入配置")
        self.setModal(True)
        self.resize(700, 600)
        self.csv_file_path = csv_file_path
        self.csv_columns = []
        # 用于记录可编辑的自定义字段名称输入框
        self.custom_name_edits = {}
        
        self.init_ui()
        self.load_configs()
        # 打开时自动选中系统默认配置并加载映射
        try:
            from database import db
            default_id = db.get_setting('default_import_config', '')
            if default_id:
                for i in range(self.config_combo.count()):
                    if str(self.config_combo.itemData(i)) == str(default_id):
                        self.config_combo.setCurrentIndex(i)
                        # 调用已有的加载方法以展示默认映射
                        self.load_config(self.config_combo.currentText())
                        break
        except Exception:
            pass
        if csv_file_path:
            self.auto_detect_encoding()
            self.load_csv_columns()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 配置选择
        config_group = QGroupBox("导入配置")
        config_layout = QGridLayout(config_group)
        
        config_layout.addWidget(QLabel("配置名称:"), 0, 0)
        self.config_combo = QComboBox()
        config_layout.addWidget(self.config_combo, 0, 1)
        
        self.new_config_btn = QPushButton("新建配置")
        self.new_config_btn.clicked.connect(self.new_config)
        config_layout.addWidget(self.new_config_btn, 0, 2)
        
        layout.addWidget(config_group)
        
        # 文件信息显示
        if self.csv_file_path:
            file_info_group = QGroupBox("文件信息")
            file_info_layout = QVBoxLayout(file_info_group)
            
            # 显示文件路径
            path_label = QLabel(f"文件路径: {self.csv_file_path}")
            path_label.setWordWrap(True)
            path_label.setStyleSheet("color: #666; font-size: 12px;")
            file_info_layout.addWidget(path_label)
            
            layout.addWidget(file_info_group)
        
        # 隐藏的编码和分隔符设置（系统自动识别）
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(['utf-8', 'gbk', 'gb2312', 'utf-8-sig'])
        self.encoding_combo.setVisible(False)
        
        self.delimiter_combo = QComboBox()
        self.delimiter_combo.addItems([',', ';', '\t', '|'])
        self.delimiter_combo.setVisible(False)
        
        # 清单列预览
        if self.csv_file_path:
            preview_group = QGroupBox("清单文件列预览")
            preview_layout = QVBoxLayout(preview_group)
            
            # 创建表格来显示清单列信息
            self.preview_table = QTableWidget()
            # 调高预览区域高度，确保完整显示列标题与示例内容
            self.preview_table.setMaximumHeight(180)
            self.preview_table.setAlternatingRowColors(True)
            self.preview_table.setSelectionBehavior(QTableWidget.SelectColumns)
            self.preview_table.horizontalHeader().setStretchLastSection(False)
            self.preview_table.verticalHeader().setVisible(False)
            # 始终显示滚动条，保障列较多时仍能完整查看
            self.preview_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
            self.preview_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            
            preview_layout.addWidget(self.preview_table)
            layout.addWidget(preview_group)
        
        # 字段映射（下方红框，支持水平滚动）
        mapping_group = QGroupBox("字段映射")
        group_vlayout = QVBoxLayout(mapping_group)
        mapping_scroll = QScrollArea()
        mapping_scroll.setWidgetResizable(True)
        mapping_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        mapping_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        mapping_container = QWidget()
        mapping_container.setMinimumWidth(300)  # 触发水平滚动，便于横向查看/编辑
        mapping_layout = QGridLayout(mapping_container)
        
        # 添加表头
        mapping_layout.addWidget(QLabel("系统字段"), 0, 0)
        mapping_layout.addWidget(QLabel("对应清单列"), 0, 1)
        
        # 系统字段列表（其中自定义字段名称可编辑）
        system_fields = [
            ('order_number', '订单号'),
            ('component_name', '板件名'),
            ('material', '材质'),
            ('finished_size', '成品尺寸'),
            ('component_code', '板件编码'),
            ('room_number', '房间号'),
            ('cabinet_number', '柜号'),
            ('remarks', '备注'),
            ('custom_field1', '自定义字段1'),
            ('custom_field2', '自定义字段2')
        ]

        self.field_combos = {}
        for i, (field_key, field_name) in enumerate(system_fields):
            row = i + 1
            if field_key in ('custom_field1', 'custom_field2'):
                # 使用可编辑的名称输入框，默认只读，双击启用编辑
                name_edit = QLineEdit(field_name)
                name_edit.setReadOnly(True)
                name_edit.setFrame(False)
                name_edit.setToolTip("双击可编辑该自定义字段的显示名称")
                name_edit.installEventFilter(self)
                # 结束编辑后恢复只读
                name_edit.editingFinished.connect(lambda ne=name_edit: ne.setReadOnly(True))
                self.custom_name_edits[field_key] = name_edit
                mapping_layout.addWidget(name_edit, row, 0)
            else:
                mapping_layout.addWidget(QLabel(f"{field_name}:"), row, 0)
            combo = QComboBox()
            combo.addItem("-- 不映射 --", "")
            self.field_combos[field_key] = combo
            mapping_layout.addWidget(combo, row, 1)
        
        mapping_scroll.setWidget(mapping_container)
        group_vlayout.addWidget(mapping_scroll)
        layout.addWidget(mapping_group)
        
        # 按钮
        button_layout = QHBoxLayout()
        self.save_btn = QPushButton("保存配置")
        self.save_btn.clicked.connect(self.save_config)
        button_layout.addWidget(self.save_btn)
        
        self.delete_btn = QPushButton("删除配置")
        self.delete_btn.clicked.connect(self.delete_config)
        button_layout.addWidget(self.delete_btn)
        
        button_layout.addStretch()
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        button_layout.addWidget(buttons)
        
        layout.addLayout(button_layout)
        
        # 连接信号
        self.config_combo.currentTextChanged.connect(self.load_config)
    
    def load_csv_columns(self):
        """加载文件的列信息（支持CSV和Excel格式）"""
        if not self.csv_file_path:
            return
            
        try:
            import os
            file_ext = os.path.splitext(self.csv_file_path)[1].lower()
            
            if file_ext == '.csv':
                headers = self._load_csv_headers()
            elif file_ext in ['.xlsx', '.xls']:
                headers = self._load_excel_headers()
            else:
                raise ValueError(f"不支持的文件格式: {file_ext}")
                
            # 更新列信息
            self.csv_columns = headers
            
            # 更新预览表格
            self._update_preview_table(headers)
            
            # 更新字段映射下拉框
            self.update_field_combos()
                
        except Exception as e:
            self._show_error_in_preview(f"读取文件失败：{str(e)}")
    
    def _load_csv_headers(self):
        """加载CSV文件的表头"""
        import csv
        encoding = self.encoding_combo.currentText()
        delimiter = self.delimiter_combo.currentText()
        
        with open(self.csv_file_path, 'r', encoding=encoding) as file:
            reader = csv.reader(file, delimiter=delimiter)
            headers = next(reader)
            return headers
    
    def _load_excel_headers(self):
        """加载Excel文件的表头"""
        try:
            import pandas as pd
            # 读取Excel文件的第一行作为表头
            df = pd.read_excel(self.csv_file_path, nrows=0)  # 只读取表头
            headers = df.columns.tolist()
            return headers
        except ImportError:
            # 如果没有pandas，尝试使用openpyxl
            try:
                from openpyxl import load_workbook
                wb = load_workbook(self.csv_file_path, read_only=True)
                ws = wb.active
                headers = []
                for cell in ws[1]:  # 第一行
                    headers.append(str(cell.value) if cell.value is not None else "")
                wb.close()
                return headers
            except ImportError:
                raise ImportError("需要安装 pandas 或 openpyxl 来支持Excel文件")
    
    def _update_preview_table(self, headers):
        """更新预览表格"""
        if hasattr(self, 'preview_table'):
            self.preview_table.setRowCount(2)  # 两行：列号和列名
            self.preview_table.setColumnCount(len(headers))
            
            # 设置表头
            self.preview_table.setVerticalHeaderLabels(['列号', '列名'])
            
            # 填充数据
            for i, header in enumerate(headers):
                # 第一行显示列号
                col_num_item = QTableWidgetItem(f"列{i+1}")
                col_num_item.setFlags(col_num_item.flags() & ~Qt.ItemIsEditable)
                self.preview_table.setItem(0, i, col_num_item)
                
                # 第二行显示列名
                col_name_item = QTableWidgetItem(str(header))
                col_name_item.setFlags(col_name_item.flags() & ~Qt.ItemIsEditable)
                self.preview_table.setItem(1, i, col_name_item)
            
            # 调整列宽
            self.preview_table.resizeColumnsToContents()
            
            # 设置水平表头为列号
            header_labels = [f"{i+1}" for i in range(len(headers))]
            self.preview_table.setHorizontalHeaderLabels(header_labels)
    
    def _show_error_in_preview(self, error_message):
        """在预览表格中显示错误信息"""
        if hasattr(self, 'preview_table'):
            self.preview_table.setRowCount(1)
            self.preview_table.setColumnCount(1)
            self.preview_table.setVerticalHeaderLabels(['错误'])
            self.preview_table.setHorizontalHeaderLabels(['信息'])
            
            error_item = QTableWidgetItem(error_message)
            error_item.setFlags(error_item.flags() & ~Qt.ItemIsEditable)
            self.preview_table.setItem(0, 0, error_item)
            self.preview_table.resizeColumnsToContents()
    
    def reload_csv_columns(self):
        """重新加载CSV列（当编码或分隔符改变时）"""
        self.load_csv_columns()
    
    def update_field_combos(self):
        """更新字段映射下拉框"""
        for combo in self.field_combos.values():
            combo.clear()
            combo.addItem("-- 不映射 --", "")
            
            # 添加列号选项
            for i, header in enumerate(self.csv_columns):
                combo.addItem(f"列{i+1}: {header}", str(i))
    
    def load_configs(self):
        """加载所有配置"""
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT config_name FROM import_configs ORDER BY is_default DESC, config_name')
        configs = cursor.fetchall()
        conn.close()
        
        self.config_combo.clear()
        for config in configs:
            self.config_combo.addItem(config[0])
    
    def load_config(self, config_name):
        """加载指定配置"""
        if not config_name:
            return
            
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT field_mapping, encoding, delimiter, custom_field_names 
            FROM import_configs WHERE config_name = ?
        ''', (config_name,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            mapping_str, encoding, delimiter, custom_names_str = result
            mapping = json.loads(mapping_str)
            
            # 设置文件参数
            self.encoding_combo.setCurrentText(encoding)
            self.delimiter_combo.setCurrentText(delimiter)
            
            # 重新加载CSV列（如果有文件的话）
            if self.csv_file_path:
                self.load_csv_columns()
            
            # 设置字段映射
            for field_key, combo in self.field_combos.items():
                if field_key in mapping:
                    # 查找对应的列号选项
                    column_index = str(mapping[field_key])  # 确保是字符串类型
                    for i in range(combo.count()):
                        if str(combo.itemData(i)) == column_index:
                            combo.setCurrentIndex(i)
                            break

            # 载入自定义字段显示名称
            try:
                if custom_names_str:
                    custom_names = json.loads(custom_names_str)
                else:
                    custom_names = {}
            except Exception:
                custom_names = {}
            for key in ('custom_field1', 'custom_field2'):
                if key in self.custom_name_edits and key in custom_names:
                    text = str(custom_names.get(key) or '')
                    if text:
                        self.custom_name_edits[key].setText(text)
    
    def new_config(self):
        """新建配置"""
        from PyQt5.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "新建配置", "配置名称:")
        if ok and name:
            self.config_combo.addItem(name)
            self.config_combo.setCurrentText(name)
    
    def save_config(self):
        """保存配置"""
        config_name = self.config_combo.currentText()
        if not config_name:
            QMessageBox.warning(self, "警告", "请输入配置名称")
            return
        
        # 收集字段映射
        mapping = {}
        for field_key, combo in self.field_combos.items():
            column_index = combo.currentData()
            if column_index:  # 不为空字符串
                mapping[field_key] = column_index

        # 收集自定义字段显示名称
        custom_names = {}
        for key, edit in self.custom_name_edits.items():
            text = edit.text().strip()
            if text:
                custom_names[key] = text
        
        # 保存到数据库
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO import_configs 
            (config_name, field_mapping, encoding, delimiter, custom_field_names, updated_at)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (
            config_name,
            json.dumps(mapping, ensure_ascii=False),
            self.encoding_combo.currentText(),
            self.delimiter_combo.currentText(),
            json.dumps(custom_names, ensure_ascii=False)
        ))
        conn.commit()
        conn.close()
        
        QMessageBox.information(self, "成功", "配置已保存")
    
    def delete_config(self):
        """删除配置"""
        config_name = self.config_combo.currentText()
        if not config_name:
            return
        
        reply = QMessageBox.question(self, "确认", f"确定要删除配置 '{config_name}' 吗？")
        if reply == QMessageBox.Yes:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute('DELETE FROM import_configs WHERE config_name = ?', (config_name,))
            conn.commit()
            conn.close()
            
            self.load_configs()
            QMessageBox.information(self, "成功", "配置已删除")
    
    def get_config(self):
        """获取当前配置"""
        config_name = self.config_combo.currentText()
        if not config_name:
            return None
        
        mapping = {}
        for field_key, combo in self.field_combos.items():
            column_index = combo.currentData()
            if column_index:  # 不为空字符串
                mapping[field_key] = int(column_index)
        # 当前自定义字段名称（仅返回，不强制使用）
        custom_names = {k: e.text().strip() for k, e in self.custom_name_edits.items() if e.text().strip()}
        
        return {
            'name': config_name,
            'mapping': mapping,
            'encoding': self.encoding_combo.currentText(),
            'delimiter': self.delimiter_combo.currentText(),
            'custom_names': custom_names
        }

    def eventFilter(self, obj, event):
        """双击开启自定义字段名称编辑"""
        if isinstance(obj, QLineEdit) and event.type() == QEvent.MouseButtonDblClick:
            obj.setReadOnly(False)
            obj.setFocus()
            obj.selectAll()
            return True
        return super().eventFilter(obj, event)
    
    def auto_detect_encoding(self):
        """自动检测文件编码和分隔符"""
        if not self.csv_file_path:
            return
        
        import os
        file_ext = os.path.splitext(self.csv_file_path)[1].lower()
        
        # Excel文件不需要检测编码和分隔符
        if file_ext in ['.xlsx', '.xls']:
            return
        
        try:
            # 检测编码
            detected_encoding = self._detect_file_encoding()
            if detected_encoding:
                # 在下拉框中查找并设置检测到的编码
                index = self.encoding_combo.findText(detected_encoding)
                if index >= 0:
                    self.encoding_combo.setCurrentIndex(index)
                else:
                    # 如果检测到的编码不在列表中，添加它
                    self.encoding_combo.addItem(detected_encoding)
                    self.encoding_combo.setCurrentText(detected_encoding)
            
            # 检测分隔符
            detected_delimiter = self._detect_file_delimiter(detected_encoding or 'utf-8')
            if detected_delimiter:
                index = self.delimiter_combo.findText(detected_delimiter)
                if index >= 0:
                    self.delimiter_combo.setCurrentIndex(index)
                    
        except Exception as e:
            print(f"文件检测失败: {e}")
    
    def _detect_file_encoding(self):
        """检测文件编码"""
        try:
            # 尝试使用chardet库检测编码
            try:
                import chardet
                with open(self.csv_file_path, 'rb') as file:
                    raw_data = file.read(10000)  # 读取前10KB用于检测
                    result = chardet.detect(raw_data)
                    if result['confidence'] > 0.7:  # 置信度大于70%
                        encoding = result['encoding'].lower()
                        # 标准化编码名称
                        if 'utf-8' in encoding:
                            return 'utf-8'
                        elif 'gb' in encoding or 'chinese' in encoding:
                            return 'gbk'
                        else:
                            return encoding
            except ImportError:
                pass
            
            # 如果没有chardet，使用简单的编码尝试
            encodings_to_try = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            
            for encoding in encodings_to_try:
                try:
                    with open(self.csv_file_path, 'r', encoding=encoding) as file:
                        file.read(1000)  # 尝试读取前1000个字符
                    return encoding
                except UnicodeDecodeError:
                    continue
            
            return 'utf-8'  # 默认返回utf-8
            
        except Exception:
            return 'utf-8'
    
    def _detect_file_delimiter(self, encoding):
        """检测文件分隔符"""
        try:
            import csv
            
            # 读取文件的前几行来检测分隔符
            with open(self.csv_file_path, 'r', encoding=encoding) as file:
                sample = file.read(1024)  # 读取前1KB
                
            # 使用csv.Sniffer来检测分隔符
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample, delimiters=',;\t|').delimiter
            
            return delimiter
            
        except Exception:
            # 如果自动检测失败，手动检测
            try:
                with open(self.csv_file_path, 'r', encoding=encoding) as file:
                    first_line = file.readline()
                    
                # 统计各种分隔符的出现次数
                delimiters = [',', ';', '\t', '|']
                counts = {}
                for delimiter in delimiters:
                    counts[delimiter] = first_line.count(delimiter)
                
                # 返回出现次数最多的分隔符
                if max(counts.values()) > 0:
                    return max(counts, key=counts.get)
                    
            except Exception:
                pass
                
            return ','  # 默认返回逗号

class ImportWorker(QThread):
    """清单导入工作线程"""
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(dict)
    error = pyqtSignal(str)
    
    def __init__(self, file_path, config, order_id, order_number, customer_name, customer_address):
        super().__init__()
        self.file_path = file_path
        self.config = config
        self.order_id = order_id
        self.order_number = order_number
        self.customer_name = customer_name
        self.customer_address = customer_address
    
    def run(self):
        try:
            import os
            file_ext = os.path.splitext(self.file_path)[1].lower()
            
            if file_ext == '.csv':
                self.status.emit("正在读取CSV文件...")
                rows = self._read_csv_file()
            elif file_ext in ['.xlsx', '.xls']:
                self.status.emit("正在读取Excel文件...")
                rows = self._read_excel_file()
            else:
                raise ValueError(f"不支持的文件格式: {file_ext}")
            
            # 跳过表头行
            if rows:
                rows = rows[1:]
            
            total_rows = len(rows)
            self.status.emit(f"共读取到 {total_rows} 行数据")
            
            # 使用现有订单
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # 使用传入的订单ID
            order_id = self.order_id
            
            # 导入板件数据
            success_count = 0
            error_count = 0
            errors = []
            
            for i, row in enumerate(rows):
                try:
                    self.progress.emit(int((i + 1) / total_rows * 100))
                    self.status.emit(f"正在导入第 {i + 1}/{total_rows} 行数据...")
                    
                    # 根据配置映射字段（使用列号）
                    component_data = {}
                    for field_key, column_index in self.config['mapping'].items():
                        if column_index < len(row):
                            component_data[field_key] = row[column_index].strip()
                    
                    # 处理component_code，保持原始编码
                    component_code = component_data.get('component_code', '').strip()
                    if not component_code:
                        # 如果component_code为空，生成一个代码（但不添加序号）
                        component_code = f"{self.order_number}_COMP_{i+1:04d}"
                    
                    # 检查component_code是否已存在，如果存在则添加后缀
                    original_code = component_code
                    suffix = 1
                    while True:
                        cursor.execute('SELECT id FROM components WHERE component_code = ?', (component_code,))
                        if not cursor.fetchone():
                            break
                        component_code = f"{original_code}_{suffix:03d}"
                        suffix += 1
                    
                    # 插入板件数据
                    cursor.execute('''
                        INSERT INTO components 
                        (order_id, component_name, material, finished_size, 
                         component_code, room_number, cabinet_number, remarks, custom_field1, custom_field2)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        order_id,
                        component_data.get('component_name', ''),
                        component_data.get('material', ''),
                        component_data.get('finished_size', ''),
                        component_code,
                        component_data.get('room_number', ''),
                        component_data.get('cabinet_number', ''),
                        component_data.get('remarks', ''),
                        component_data.get('custom_field1', ''),
                        component_data.get('custom_field2', '')
                    ))
                    
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"第{i+1}行: {str(e)}")
            
            conn.commit()
            conn.close()
            
            # 记录操作日志
            db.log_operation('import_csv', {
                'order_id': order_id,
                'file_path': self.file_path,
                'total_rows': total_rows,
                'success_count': success_count,
                'error_count': error_count
            })
            
            result = {
                'order_id': order_id,
                'total_rows': total_rows,
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            
            self.finished.emit(result)
            
        except Exception as e:
            self.error.emit(str(e))
    
    def _read_csv_file(self):
        """读取CSV文件"""
        import csv
        with open(self.file_path, 'r', encoding=self.config['encoding']) as file:
            # 检测分隔符
            delimiter = self.config['delimiter']
            if delimiter == '\\t':
                delimiter = '\t'
            
            reader = csv.reader(file, delimiter=delimiter)
            rows = list(reader)
        return rows
    
    def _read_excel_file(self):
        """读取Excel文件"""
        try:
            import pandas as pd
            # 读取Excel文件
            df = pd.read_excel(self.file_path, header=None)
            # 转换为列表格式
            rows = df.values.tolist()
            # 将NaN值转换为空字符串
            for i, row in enumerate(rows):
                rows[i] = [str(cell) if pd.notna(cell) else '' for cell in row]
            return rows
        except ImportError:
            # 如果没有pandas，尝试使用openpyxl
            try:
                from openpyxl import load_workbook
                wb = load_workbook(self.file_path, read_only=True)
                ws = wb.active
                rows = []
                for row in ws.iter_rows(values_only=True):
                    # 将None值转换为空字符串
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    rows.append(row_data)
                wb.close()
                return rows
            except ImportError:
                raise ImportError("需要安装 pandas 或 openpyxl 来支持Excel文件")

class OrderManagement(QWidget):
    """订单管理模块"""
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_orders()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 工具栏
        toolbar_layout = QHBoxLayout()
        
        self.new_order_btn = QPushButton("新建订单")
        self.new_order_btn.clicked.connect(self.new_order)
        toolbar_layout.addWidget(self.new_order_btn)
        
        self.edit_order_btn = QPushButton("编辑订单")
        self.edit_order_btn.clicked.connect(self.edit_order)
        self.edit_order_btn.setEnabled(False)  # 默认禁用，需要选择订单后启用
        toolbar_layout.addWidget(self.edit_order_btn)
        
        self.delete_order_btn = QPushButton("删除订单")
        self.delete_order_btn.clicked.connect(self.delete_order)
        self.delete_order_btn.setEnabled(False)  # 默认禁用，需要选择订单后启用
        toolbar_layout.addWidget(self.delete_order_btn)

        self.import_csv_btn = QPushButton("导入CSV")
        self.import_csv_btn.clicked.connect(self.import_csv_data)
        toolbar_layout.addWidget(self.import_csv_btn)

        self.config_import_btn = QPushButton("导入配置")
        self.config_import_btn.clicked.connect(self.config_import)
        toolbar_layout.addWidget(self.config_import_btn)

        toolbar_layout.addStretch()
        
        self.refresh_btn = QPushButton("刷新")
        self.refresh_btn.clicked.connect(self.load_orders)
        toolbar_layout.addWidget(self.refresh_btn)
        
        layout.addLayout(toolbar_layout)
        
        # 分割器
        splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(splitter)
        
        # 左侧：订单列表
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        
        left_layout.addWidget(QLabel("订单列表"))
        
        # 搜索框布局
        search_layout = QHBoxLayout()
        
        # 订单号搜索
        search_layout.addWidget(QLabel("订单号:"))
        self.order_number_search = QLineEdit()
        self.order_number_search.setPlaceholderText("输入订单号搜索...")
        self.order_number_search.textChanged.connect(self.filter_orders)
        search_layout.addWidget(self.order_number_search)
        
        # 客户地址搜索
        search_layout.addWidget(QLabel("客户地址:"))
        self.customer_address_search = QLineEdit()
        self.customer_address_search.setPlaceholderText("输入客户地址搜索...")
        self.customer_address_search.textChanged.connect(self.filter_orders)
        search_layout.addWidget(self.customer_address_search)
        
        left_layout.addLayout(search_layout)
        
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(5)
        self.orders_table.setHorizontalHeaderLabels(['订单号', '客户名称', '客户地址', '创建时间', '状态'])
        self.orders_table.horizontalHeader().setStretchLastSection(True)
        self.orders_table.selectionBehavior = QTableWidget.SelectRows
        self.orders_table.itemSelectionChanged.connect(self.on_order_selected)
        
        # 设置滚动条策略
        self.orders_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.orders_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        left_layout.addWidget(self.orders_table)
        
        splitter.addWidget(left_widget)
        
        # 右侧：订单详情
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        
        # 订单信息
        order_info_group = QGroupBox("订单信息")
        order_info_layout = QGridLayout(order_info_group)
        
        order_info_layout.addWidget(QLabel("订单号:"), 0, 0)
        self.order_number_label = QLabel("-")
        order_info_layout.addWidget(self.order_number_label, 0, 1)
        
        order_info_layout.addWidget(QLabel("客户名称:"), 1, 0)
        self.customer_name_label = QLabel("-")
        order_info_layout.addWidget(self.customer_name_label, 1, 1)
        
        order_info_layout.addWidget(QLabel("客户地址:"), 2, 0)
        self.customer_address_label = QLabel("-")
        order_info_layout.addWidget(self.customer_address_label, 2, 1)
        
        order_info_layout.addWidget(QLabel("创建时间:"), 3, 0)
        self.created_at_label = QLabel("-")
        order_info_layout.addWidget(self.created_at_label, 3, 1)
        
        right_layout.addWidget(order_info_group)
        
        # 板件列表
        components_group = QGroupBox("板件列表")
        components_layout = QVBoxLayout(components_group)
        
        self.components_table = QTableWidget()
        self.components_table.setColumnCount(7)
        self.components_table.setHorizontalHeaderLabels([
            '板件名', '材质', '成品尺寸', '板件编码', '房间号', '柜号', '状态'
        ])
        self.components_table.horizontalHeader().setStretchLastSection(True)
        components_layout.addWidget(self.components_table)
        
        right_layout.addWidget(components_group)
        
        splitter.addWidget(right_widget)
        
        # 设置分割器比例
        splitter.setSizes([400, 600])
    
    def load_orders(self):
        """加载订单列表"""
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, order_number, customer_name, customer_address, 
                   created_at, status
            FROM orders ORDER BY created_at DESC
        ''')
        orders = cursor.fetchall()
        conn.close()
        
        self.orders_table.setRowCount(len(orders))
        for i, order in enumerate(orders):
            self.orders_table.setItem(i, 0, QTableWidgetItem(order[1]))  # order_number
            self.orders_table.setItem(i, 1, QTableWidgetItem(order[2] or ''))  # customer_name
            self.orders_table.setItem(i, 2, QTableWidgetItem(order[3] or ''))  # customer_address
            self.orders_table.setItem(i, 3, QTableWidgetItem(order[4]))  # created_at
            # 中文映射订单状态
            status_map = {
                'active': '活跃',
                'inactive': '停用',
                'archived': '已归档',
                'closed': '已关闭',
                'open': '进行中',
            }
            status_text = status_map.get(str(order[5]), str(order[5]) if order[5] else '未设置')
            self.orders_table.setItem(i, 4, QTableWidgetItem(status_text))  # status
            
            # 存储订单ID
            self.orders_table.item(i, 0).setData(Qt.UserRole, order[0])
    
    def on_order_selected(self):
        """订单选择事件"""
        current_row = self.orders_table.currentRow()
        if current_row >= 0:
            order_id = self.orders_table.item(current_row, 0).data(Qt.UserRole)
            self.load_order_details(order_id)
            # 启用编辑和删除按钮
            self.edit_order_btn.setEnabled(True)
            self.delete_order_btn.setEnabled(True)
        else:
            # 禁用编辑和删除按钮
            self.edit_order_btn.setEnabled(False)
            self.delete_order_btn.setEnabled(False)
    
    def load_order_details(self, order_id):
        """加载订单详情"""
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # 加载订单信息
        cursor.execute('''
            SELECT order_number, customer_name, customer_address, created_at
            FROM orders WHERE id = ?
        ''', (order_id,))
        order = cursor.fetchone()
        
        if order:
            self.order_number_label.setText(order[0])
            self.customer_name_label.setText(order[1] or '')
            self.customer_address_label.setText(order[2] or '')
            self.created_at_label.setText(order[3])
        
        # 加载板件列表
        cursor.execute('''
            SELECT component_name, material, finished_size, component_code,
                   room_number, cabinet_number, status
            FROM components WHERE order_id = ?
            ORDER BY component_name
        ''', (order_id,))
        components = cursor.fetchall()
        
        self.components_table.setRowCount(len(components))
        for i, component in enumerate(components):
            for j, value in enumerate(component):
                self.components_table.setItem(i, j, QTableWidgetItem(value or ''))
        
        conn.close()
    
    def filter_orders(self):
        """根据搜索条件过滤订单列表"""
        order_number_filter = self.order_number_search.text().lower()
        address_filter = self.customer_address_search.text().lower()
        
        for row in range(self.orders_table.rowCount()):
            # 获取当前行的订单号和客户地址
            order_number_item = self.orders_table.item(row, 0)
            address_item = self.orders_table.item(row, 2)
            
            order_number = order_number_item.text().lower() if order_number_item else ""
            address = address_item.text().lower() if address_item else ""
            
            # 检查是否匹配搜索条件
            order_match = order_number_filter == "" or order_number_filter in order_number
            address_match = address_filter == "" or address_filter in address
            
            # 显示或隐藏行
            self.orders_table.setRowHidden(row, not (order_match and address_match))
    
    def new_order(self):
        """新建订单"""
        dialog = NewOrderDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            order_data = dialog.get_order_data()
            order_id = order_data['id']
            
            # 重新加载订单列表
            self.load_orders()
            
            # 自动选择新创建的订单
            for i in range(self.orders_table.rowCount()):
                if self.orders_table.item(i, 0).data(Qt.UserRole) == order_id:
                    self.orders_table.selectRow(i)
                    break
            
            QMessageBox.information(self, "成功", "订单创建成功")
    
    def config_import(self):
        """配置导入"""
        # 选择文件
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", "", 
            "所有支持的文件 (*.csv *.xlsx *.xls);;CSV文件 (*.csv);;Excel文件 (*.xlsx *.xls)")
        
        if not file_path:
            return
        
        # 打开配置对话框
        dialog = ImportConfigDialog(self, file_path)
        dialog.exec_()
    
    def import_csv_data(self):
        """导入CSV数据"""
        # 首先选择订单
        order_dialog = OrderSelectionDialog(self)
        if order_dialog.exec_() != QDialog.Accepted:
            return
        
        selected_order = order_dialog.get_selected_order()
        if not selected_order:
            QMessageBox.warning(self, "警告", "请选择一个订单")
            return
        
        # 选择文件
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据文件", "", 
            "所有支持的文件 (*.csv *.xlsx *.xls);;CSV文件 (*.csv);;Excel文件 (*.xlsx *.xls)")
        
        if not file_path:
            return
        
        # 配置导入（传递CSV文件路径）
        config_dialog = ImportConfigDialog(self, file_path)
        if config_dialog.exec_() != QDialog.Accepted:
            return
        
        config = config_dialog.get_config()
        if not config:
            QMessageBox.warning(self, "警告", "请配置字段映射")
            return
        
        # 使用选中的订单信息
        order_number = selected_order['order_number']
        customer_name = selected_order['customer_name']
        customer_address = selected_order['customer_address']
        
        # 显示进度对话框
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("导入进度")
        progress_dialog.setModal(True)
        progress_dialog.resize(400, 150)
        
        progress_layout = QVBoxLayout(progress_dialog)
        
        status_label = QLabel("准备导入...")
        progress_layout.addWidget(status_label)
        
        progress_bar = QProgressBar()
        progress_layout.addWidget(progress_bar)
        
        cancel_btn = QPushButton("取消")
        progress_layout.addWidget(cancel_btn)
        
        # 启动导入线程
        self.import_worker = ImportWorker(file_path, config, selected_order['id'], order_number, customer_name, customer_address)
        self.import_worker.progress.connect(progress_bar.setValue)
        self.import_worker.status.connect(status_label.setText)
        self.import_worker.finished.connect(lambda result: self.import_finished(result, progress_dialog))
        self.import_worker.error.connect(lambda error: self.import_error(error, progress_dialog))
        
        cancel_btn.clicked.connect(lambda: self.cancel_import(progress_dialog))
        
        self.import_worker.start()
        progress_dialog.exec_()
    
    def import_finished(self, result, progress_dialog):
        """导入完成"""
        progress_dialog.accept()
        
        message = f"""导入完成！
        
总行数: {result['total_rows']}
成功导入: {result['success_count']}
失败: {result['error_count']}"""
        
        if result['errors']:
            message += f"\n\n错误详情:\n" + "\n".join(result['errors'][:10])
            if len(result['errors']) > 10:
                message += f"\n... 还有 {len(result['errors']) - 10} 个错误"
        
        QMessageBox.information(self, "导入结果", message)
        self.load_orders()
    
    def import_error(self, error, progress_dialog):
        """导入错误"""
        progress_dialog.reject()
        QMessageBox.critical(self, "导入错误", f"导入失败：\n{str(error)}")
    
    def delete_order(self):
        """删除订单"""
        current_row = self.orders_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请先选择要删除的订单")
            return
        
        # 获取订单信息
        order_id = self.orders_table.item(current_row, 0).data(Qt.UserRole)
        order_number = self.orders_table.item(current_row, 0).text()
        customer_name = self.orders_table.item(current_row, 1).text()
        
        # 确认删除
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除订单吗？\n\n订单号: {order_number}\n客户: {customer_name}\n\n注意：删除订单将同时删除该订单下的所有板件数据，此操作不可恢复！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply != QMessageBox.Yes:
            return
        
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            # 业务校验：若订单下存在已入包/入托板件或包裹，则禁止删除
            cursor.execute('SELECT COUNT(*) FROM components WHERE order_id = ? AND package_id IS NOT NULL', (order_id,))
            packed_cnt = cursor.fetchone()[0]
            cursor.execute('SELECT COUNT(*) FROM packages WHERE order_id = ? AND pallet_id IS NOT NULL', (order_id,))
            pallet_cnt = cursor.fetchone()[0]
            if packed_cnt > 0 or pallet_cnt > 0:
                conn.close()
                QMessageBox.warning(self, "警告", "该订单包含已入包或已入托的板件/包裹，不能删除。请先在打包/托盘模块处理后再尝试。")
                return

            # 删除订单相关的未入包板件数据
            cursor.execute('DELETE FROM components WHERE order_id = ? AND package_id IS NULL', (order_id,))
            
            # 删除订单
            cursor.execute('DELETE FROM orders WHERE id = ?', (order_id,))
            
            conn.commit()
            conn.close()
            
            # 记录操作日志
            db.log_operation('delete_order', {
                'order_id': order_id,
                'order_number': order_number,
                'customer_name': customer_name
            })
            
            QMessageBox.information(self, "成功", f"订单 {order_number} 已成功删除")
            
            # 刷新订单列表
            self.load_orders()
            
            # 清空右侧详情显示
            self.order_number_label.setText("-")
            self.customer_name_label.setText("-")
            self.customer_address_label.setText("-")
            self.created_at_label.setText("-")
            self.components_table.setRowCount(0)
            
        except Exception as e:
             QMessageBox.critical(self, "错误", f"删除订单失败：\n{str(e)}")
    
    def edit_order(self):
        """编辑订单"""
        current_row = self.orders_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请先选择要编辑的订单")
            return
        
        # 获取订单信息
        order_id = self.orders_table.item(current_row, 0).data(Qt.UserRole)
        
        # 创建编辑对话框
        dialog = EditOrderDialog(self, order_id)
        if dialog.exec_() == QDialog.Accepted:
            # 刷新订单列表
            self.load_orders()
            # 重新选择当前订单
            for row in range(self.orders_table.rowCount()):
                if self.orders_table.item(row, 0).data(Qt.UserRole) == order_id:
                    self.orders_table.selectRow(row)
                    break
    
    def cancel_import(self, progress_dialog):
        """取消导入"""
        if hasattr(self, 'import_worker'):
            self.import_worker.terminate()
        progress_dialog.reject()